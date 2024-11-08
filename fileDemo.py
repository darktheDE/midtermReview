# --- Hướng dẫn cơ bản về Nhập xuất file trong Python ---

# ===================== 1. Mở và đóng file =====================

# Trong Python, để làm việc với file, chúng ta cần mở file bằng hàm `open()` và đóng file bằng hàm `close()`.
# Hàm `open()` nhận hai đối số: tên file và chế độ mở file.

# Các chế độ mở file thông dụng:
# 'r'  : Đọc file (mặc định), báo lỗi nếu file không tồn tại.
# 'w'  : Ghi file, tạo file mới nếu file không tồn tại, ghi đè nếu file đã tồn tại.
# 'a'  : Ghi tiếp vào cuối file mà không ghi đè, tạo file mới nếu file không tồn tại.
# 'r+' : Đọc và ghi, báo lỗi nếu file không tồn tại.
# 'b'  : Mở file ở chế độ nhị phân (binary) (vd. 'rb', 'wb').

# a) Mở và đóng file
file = open("example.txt", "w")  # Mở file để ghi
file.write("Hello, Python!")     # Ghi nội dung vào file
file.close()                     # Đóng file

# Lưu ý: Luôn đóng file sau khi hoàn thành để giải phóng tài nguyên hệ thống.

# ===================== 2. Ghi file =====================

# a) Ghi vào file văn bản bằng `write()`
with open("example.txt", "w") as file:  # Mở file ở chế độ ghi
    file.write("Dòng đầu tiên.\n")      # Ghi một dòng và xuống dòng
    file.write("Dòng thứ hai.\n")       # Ghi thêm một dòng khác
    # File sẽ tự động đóng sau khi thoát khỏi khối `with`

# b) Ghi nhiều dòng vào file với `writelines()`
lines = ["Dòng thứ ba.\n", "Dòng thứ tư.\n"]
with open("example.txt", "a") as file:  # Mở file ở chế độ ghi tiếp
    file.writelines(lines)

# ===================== 3. Đọc file =====================

# a) Đọc toàn bộ nội dung của file bằng `read()`
with open("example.txt", "r") as file:
    content = file.read()              # Đọc toàn bộ nội dung file
    print("Nội dung của file:\n", content)

# b) Đọc từng dòng một với `readline()`
with open("example.txt", "r") as file:
    line1 = file.readline()            # Đọc dòng đầu tiên
    line2 = file.readline()            # Đọc dòng thứ hai
    print("Dòng 1:", line1)
    print("Dòng 2:", line2)

# c) Đọc tất cả các dòng vào một list với `readlines()`
with open("example.txt", "r") as file:
    lines = file.readlines()           # Đọc tất cả các dòng vào list
    print("Các dòng trong file:", lines)

# ===================== 4. Xử lý từng dòng trong file =====================

# Trong một số trường hợp, bạn có thể muốn xử lý từng dòng của file để tiết kiệm bộ nhớ,
# đặc biệt khi file có dung lượng lớn.

with open("example.txt", "r") as file:
    for line in file:
        print("Xử lý dòng:", line.strip())  # Xử lý dòng sau khi loại bỏ khoảng trắng và ký tự xuống dòng

# ===================== 5. Làm việc với file nhị phân =====================

# Khi làm việc với các file nhị phân (vd. file ảnh, video), bạn cần mở file ở chế độ nhị phân ('b').

# a) Ghi dữ liệu nhị phân vào file
binary_data = b'\x00\xFF\xAC\x01'  # Một chuỗi dữ liệu nhị phân
with open("binary_file.bin", "wb") as binary_file:
    binary_file.write(binary_data)

# b) Đọc dữ liệu nhị phân từ file
with open("binary_file.bin", "rb") as binary_file:
    binary_content = binary_file.read()
    print("Dữ liệu nhị phân:", binary_content)

# ===================== 6. Xóa và đổi tên file =====================

# Để xóa hoặc đổi tên file, bạn cần sử dụng module `os`.
import os

# a) Xóa file
# os.remove("example.txt")  # Xóa file example.txt

# b) Đổi tên file
# os.rename("example.txt", "new_example.txt")  # Đổi tên example.txt thành new_example.txt

# ===================== 7. Các phương pháp xử lý file an toàn với `with` =====================

# Trong Python, khi làm việc với file, bạn nên dùng khối `with` để đảm bảo file được đóng tự động sau khi hoàn thành.
# Điều này giúp tránh lỗi nếu quên đóng file.

# Ví dụ:
with open("example.txt", "r") as file:
    content = file.read()
    print("Nội dung file (sử dụng with):", content)

# ===================== 8. Lưu trữ dữ liệu vào file CSV =====================

# CSV (Comma-Separated Values) là định dạng file phổ biến để lưu trữ dữ liệu dạng bảng.

import csv

# a) Ghi vào file CSV
data = [["Tên", "Tuổi", "Thành phố"],
        ["Alice", 25, "New York"],
        ["Bob", 30, "San Francisco"]]

with open("people.csv", "w", newline='') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerows(data)  # Ghi tất cả các dòng dữ liệu vào file CSV

# b) Đọc file CSV
with open("people.csv", "r") as csvfile:
    reader = csv.reader(csvfile)
    for row in reader:
        print("Dòng CSV:", row)

# ===================== 9. Đọc và ghi file JSON =====================

# JSON (JavaScript Object Notation) là định dạng dữ liệu phổ biến cho các ứng dụng web và API.

import json

# a) Ghi dữ liệu vào file JSON
data = {
    "name": "Alice",
    "age": 25,
    "city": "New York"
}

with open("data.json", "w") as jsonfile:
    json.dump(data, jsonfile)  # Ghi dữ liệu JSON vào file

# b) Đọc dữ liệu từ file JSON
with open("data.json", "r") as jsonfile:
    data = json.load(jsonfile)  # Đọc dữ liệu JSON từ file
    print("Dữ liệu JSON:", data)

# ===================== 10. Làm việc với các file lớn =====================

# Khi làm việc với các file lớn, đọc file theo từng dòng hoặc từng phần nhỏ giúp tiết kiệm bộ nhớ.

# Ví dụ đọc file lớn từng dòng một:
with open("large_file.txt", "r") as file:
    for line in file:
        process_line = line.strip()  # Xử lý từng dòng
        print("Dòng xử lý:", process_line)

# --- Kết thúc hướng dẫn cơ bản về nhập xuất file trong Python ---
# Các thao tác trên cung cấp cho bạn kiến thức cần thiết để làm việc với file văn bản, nhị phân, CSV và JSON,
# bao gồm các phương pháp ghi, đọc, xử lý và quản lý file.
# --- Hướng dẫn bổ sung về Nhập xuất file trong Python ---

# ===================== 3. Xử lý tập tin XML =====================

# XML (eXtensible Markup Language) là định dạng phổ biến để lưu trữ và trao đổi dữ liệu.
# Để xử lý file XML, Python cung cấp module `xml.etree.ElementTree`.

import xml.etree.ElementTree as ET

# a) Đọc file XML
tree = ET.parse("example.xml")          # Phân tích cú pháp file XML
root = tree.getroot()                    # Lấy phần tử gốc của file XML

# Duyệt qua các phần tử trong file XML
for child in root:
    print("Tag:", child.tag, ", Attribute:", child.attrib)
    for subchild in child:
        print("   Tag:", subchild.tag, ", Text:", subchild.text)

# b) Ghi dữ liệu vào file XML
# Tạo cấu trúc XML từ các phần tử
data = ET.Element("data")
person = ET.SubElement(data, "person", name="Alice")
age = ET.SubElement(person, "age")
age.text = "25"

# Ghi cấu trúc XML vào file
tree = ET.ElementTree(data)
tree.write("output.xml", encoding="utf-8", xml_declaration=True)

# c) Thay đổi hoặc thêm phần tử XML
# Thêm một phần tử mới vào file XML hiện có
new_person = ET.SubElement(root, "person", name="Bob")
new_age = ET.SubElement(new_person, "age")
new_age.text = "30"

tree.write("updated_example.xml", encoding="utf-8", xml_declaration=True)

# ===================== 6. Xử lý tập tin Excel =====================

# Excel là một định dạng file phổ biến để lưu trữ dữ liệu bảng.
# Để xử lý file Excel trong Python, chúng ta thường sử dụng thư viện `openpyxl` hoặc `pandas`.

# Cài đặt thư viện `openpyxl` và `pandas` nếu chưa có:
# pip install openpyxl pandas

import openpyxl
import pandas as pd

# a) Đọc file Excel bằng `openpyxl`
workbook = openpyxl.load_workbook("example.xlsx")  # Mở file Excel
sheet = workbook.active                            # Lấy trang tính (sheet) đầu tiên

# Duyệt qua các ô (cells) trong sheet
for row in sheet.iter_rows(values_only=True):
    print("Dòng:", row)

# b) Ghi dữ liệu vào file Excel bằng `openpyxl`
workbook = openpyxl.Workbook()               # Tạo workbook mới
sheet = workbook.active                      # Lấy trang tính đầu tiên
sheet.title = "DataSheet"                    # Đặt tên cho trang tính

# Thêm dữ liệu vào các ô
sheet["A1"] = "Tên"
sheet["B1"] = "Tuổi"
sheet.append(["Alice", 25])
sheet.append(["Bob", 30])

# Lưu workbook vào file Excel
workbook.save("output.xlsx")

# c) Đọc và ghi file Excel bằng `pandas`
# `pandas` cung cấp các hàm tiện lợi để xử lý file Excel dưới dạng DataFrame.

# Đọc file Excel vào DataFrame
df = pd.read_excel("example.xlsx", sheet_name="Sheet1")
print("Dữ liệu từ file Excel:\n", df)

# Ghi DataFrame vào file Excel
data = {"Tên": ["Alice", "Bob"], "Tuổi": [25, 30]}
df = pd.DataFrame(data)
df.to_excel("output_pandas.xlsx", index=False)

# --- Kết thúc hướng dẫn bổ sung về xử lý tập tin XML và Excel ---
# Với các bước trên, bạn đã có thêm kiến thức để làm việc với file XML và Excel trong Python.
# Các ví dụ bao gồm đọc, ghi và chỉnh sửa dữ liệu trong cả hai loại file này.
